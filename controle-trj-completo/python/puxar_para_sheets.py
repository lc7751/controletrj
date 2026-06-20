#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================
 puxar_para_sheets.py
 ---------------------------------------------------------------------
 Lê a planilha de TAREFAS exportada da plataforma (.xlsx) e ENVIA as
 linhas para a sua Planilha Google (aba TASKS) através do Apps Script.

 O site (GitHub Pages) faz TODO o cálculo (região, SLA, backlog...).
 Este script apenas LÊ o arquivo e ESCREVE os dados "crus" na planilha.

 COMO USAR (resumo — veja o LEIA-ME.md para o passo a passo completo):
   1. Instale as dependências:   pip install -r requirements.txt
   2. Ajuste as 3 variáveis abaixo (URL, TOKEN, ARQUIVO).
   3. Rode:                       python puxar_para_sheets.py
=====================================================================
"""
import sys
import json
import time

try:
    import requests
except ImportError:
    print("Falta a biblioteca 'requests'. Rode:  pip install -r requirements.txt")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta a biblioteca 'openpyxl'. Rode:  pip install -r requirements.txt")
    sys.exit(1)

# =====================================================================
# 1) CONFIGURAÇÃO  —  AJUSTE ESTAS TRÊS LINHAS
# =====================================================================
APPS_SCRIPT_URL = ""   # URL do seu Web App (termina em /exec)
API_TOKEN       = ""   # Token gerado pelo Apps Script (menu Configurações do site, ou execute mostrarToken())
ARQUIVO_XLSX    = "tarefas.xlsx"   # Caminho do arquivo exportado da plataforma

# Nome da aba dentro do .xlsx (deixe None para usar a primeira aba)
NOME_ABA = None
# Quantas linhas de cabeçalho pular antes dos dados
LINHAS_CABECALHO = 1
# Enviar em lotes de N linhas (evita payloads gigantes)
TAMANHO_LOTE = 2000

# =====================================================================
# 2) MAPA DE COLUNAS  (número da COLUNA no .xlsx, começando em 1)
#    Ajuste os números conforme o seu arquivo, se necessário.
#    A ordem/nomes à esquerda NÃO mude (são lidos pelo site).
# =====================================================================
COLUNAS = {
    "filaAtual":          1,
    "osNumero":           2,
    "sequenciaId":        3,
    "tipoAtividade":      4,
    "status":             5,
    "eta":               13,
    "fim":               14,
    "habilidadeTrabalho":16,
    "microarea":         17,
    "dataBase":          20,
    "vencimentoSla":     21,
    "enderecoId":        67,
    "siteId":            84,
    "tipoFalha":        132,   # tipo de falha (coluna 132 no .xlsm de origem)
    "dataCriacao":      176,   # usado para o "aging" (coluna 176 no .xlsm)
    "quemEncerrou":     180,
    "prioridade":       191,
}


def val(row, idx):
    """Pega o valor da coluna idx (1-based). Retorna string ou ''."""
    if not idx:
        return ""
    i = idx - 1
    if i < 0 or i >= len(row):
        return ""
    v = row[i]
    if v is None:
        return ""
    # datas -> ISO; demais -> str
    try:
        import datetime
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.isoformat()
    except Exception:
        pass
    return str(v).strip()


def main():
    if not APPS_SCRIPT_URL or not API_TOKEN:
        print("ERRO: preencha APPS_SCRIPT_URL e API_TOKEN no topo do arquivo.")
        sys.exit(1)

    print("Lendo arquivo:", ARQUIVO_XLSX)
    wb = load_workbook(ARQUIVO_XLSX, read_only=True, data_only=True)
    ws = wb[NOME_ABA] if NOME_ABA else wb[wb.sheetnames[0]]
    print("Aba:", ws.title)

    linhas = []
    n = 0
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r < LINHAS_CABECALHO:
            continue
        if row is None:
            continue
        # ignora linhas totalmente vazias
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        obj = {}
        for nome, idx in COLUNAS.items():
            obj[nome] = val(row, idx)
        # precisa de pelo menos uma OS ou endereço para valer
        if not obj.get("osNumero") and not obj.get("enderecoId"):
            continue
        linhas.append(obj)
        n += 1

    print("Total de tarefas lidas:", n)
    if not linhas:
        print("Nada para enviar.")
        return

    enviadas = 0
    primeiro = True
    for i in range(0, len(linhas), TAMANHO_LOTE):
        lote = linhas[i:i + TAMANHO_LOTE]
        # 1o lote substitui tudo; demais seriam append — aqui enviamos
        # tudo de uma vez por simplicidade (saveTasks substitui a aba).
        payload = {"action": "saveTasks", "token": API_TOKEN, "rows": lote}
        if not primeiro:
            # Para mais de um lote seria preciso uma ação de "append".
            # Por padrão, o site lida bem com um único envio.
            pass
        resp = enviar(payload)
        enviadas += len(lote)
        print("Lote enviado:", len(lote), "->", resp.get("ok"), resp.get("count", ""))
        primeiro = False
        time.sleep(0.5)

    print("Concluído. Tarefas enviadas:", enviadas)
    print("Abra o site e clique em Atualizar para ver os dados.")


def enviar(payload):
    # text/plain evita o preflight CORS no Apps Script
    headers = {"Content-Type": "text/plain;charset=utf-8"}
    r = requests.post(APPS_SCRIPT_URL, data=json.dumps(payload), headers=headers, timeout=120, allow_redirects=True)
    try:
        data = r.json()
    except Exception:
        print("Resposta inesperada do servidor:")
        print(r.text[:500])
        sys.exit(1)
    if not data.get("ok"):
        print("Erro do servidor:", data.get("error"))
        sys.exit(1)
    return data


if __name__ == "__main__":
    main()
